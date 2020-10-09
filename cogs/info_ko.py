import time
import discord
import psutil
import os
import matplotlib.pyplot as plt
import openpyxl
import asyncio

from datetime import datetime
from discord.ext import commands
from evs import default, permissions


cachelib = "./lib/cache/"


class Information_ko(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = default.get("config.json")
        self.process = psutil.Process(os.getpid())
        #파일생성
        if os.path.isfile(cachelib + "usage.xlsx"):
            print("cache file exist")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "1"
            for i in range(1, 101):
                ws.cell(row=2, column=i).value = "0"
                ws.cell(row=3, column=i).value = "0"
            wb.save(cachelib + "usage.xlsx")
            wb.close()

    @commands.command()
    async def 핑(self, ctx):
        """ Pong! """
        before = time.monotonic()
        before_ws = int(round(self.bot.latency * 1000, 1))
        message = await ctx.send("🏓 퐁")
        ping = (time.monotonic() - before) * 1000
        await message.edit(content=f"🏓 WS: {before_ws}ms  |  REST: {int(ping)}ms")

    @commands.command(aliases=['봇 초대하기', '초대', '참가'])
    async def 초대하기(self, ctx):
        """ Invite me to your server """
        embed = discord.Embed(title="저를 파티에 초대해주세요!", description=f"**{ctx.author.name}**, 아래의 링크를 사용하세요\n[link](https://discord.com/oauth2/authorize?client_id=749629426777456691&permissions=8&scope=bot)", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command()
    async def 소스코드(self, ctx):
        """ Check out my source code <3 """
        await ctx.send(f"**{ctx.bot.user}** 이 소스 코드로 돌아갑니다:\nhttps://github.com/Shio7/Keter")

    @commands.command(aliases=['지원 서버', '문의 서버'])
    async def 서버(self, ctx):
        """ Get an invite to our support server! """
        if isinstance(ctx.channel, discord.DMChannel) or ctx.guild.id != 749595288280498188:
            return await ctx.send(f"**여기로! {ctx.author.name} 🍻\n<{self.config.botserver}>**")

        await ctx.send(f"**{ctx.author.name}** 이게 제 집이잖아요~ :3")

    @commands.command(aliases=['상태'])
    async def 정보(self, ctx):
        """ About the bot """
        f = open("./lib/cache/version.ccf", "r")
        version = f.read()
        f.close
        ramUsage = self.process.memory_full_info().rss / 1024**2
        avgmembers = round(len(self.bot.users) / len(self.bot.guilds))

        embed = discord.Embed(colour=0xeff0f1)
        embed.set_thumbnail(url=ctx.bot.user.avatar_url)
        embed.add_field(name="마지막 부팅", value=default.timeago(datetime.now() - self.bot.uptime), inline=True)
        embed.add_field(
            name=f"개발자{'' if len(self.config.owners) == 1 else 's'}",
            value=', '.join([str(self.bot.get_user(x)) for x in self.config.owners]),
            inline=True)
        embed.add_field(name="라이브러리", value="discord.py", inline=True)
        embed.add_field(name="유저", value= str(len(ctx.bot.guilds)*avgmembers) + " users", inline=True)
        embed.add_field(name="커맨드 로드", value=len([x.name for x in self.bot.commands]), inline=True)
        embed.add_field(name="램", value=f"{ramUsage:.2f} MB", inline=True)

        await ctx.send(content=f"ℹ About **{ctx.bot.user}** | **" + version + "**", embed=embed)

    @commands.command(aliases=['상태기록'])
    @commands.check(permissions.is_owner)
    async def 정보기록(self, ctx):
        """ Write the status of system """
        if os.path.isfile(cachelib + "usagecheck.ccf"):
            return await ctx.send("이미 실행중입니다.")
        await ctx.send("코드를 실행합니다.")
        f = open(cachelib + "usagecheck.ccf","w")
        f.close
        while True:
            if not os.path.isfile(cachelib + "usagecheck.ccf"):
                return await ctx.send("The cache file went wrong.")
            wb = openpyxl.load_workbook(cachelib + "usage.xlsx")
            ws = wb.active
            rampercent = psutil.virtual_memory().available/(self.process.memory_full_info().rss + psutil.virtual_memory().available)
            cpupercent = self.process.cpu_percent()
            last = ws.cell(row=1, column=1).value
            if last == "100":
                ws.cell(row=2, column=1).value = str(rampercent)
                ws.cell(row=3, column=int(last) + 1).value = str(rampercent)
                ws.cell(row=1, column=1).value = "1"
            else:
                ws.cell(row=2, column=int(last) + 1).value = str(rampercent)
                ws.cell(row=3, column=int(last) + 1).value = str(rampercent)
                ws.cell(row=1, column=1).value = str(int(last) + 1)
            wb.save(cachelib + "usage.xlsx")
            wb.close()
            await asyncio.sleep(60)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 기록중지(self, ctx):
        if os.path.isfile(cachelib + "usagecheck.ccf"):
            os.remove(cachelib + "usagecheck.ccf")
            return await ctx.send("캐시가 삭제되었습니다.")
        await ctx.send("캐시가 없습니다.")
    
    @commands.command(aliases=['상태그래프'])
    @commands.check(permissions.is_owner)
    async def 정보그래프(self, ctx):
        wb = openpyxl.load_workbook(cachelib + "usage.xlsx")
        ws = wb.active
        last = ws.cell(row=1, column=1).value
        rams = []
        cpus = []
        if last == "100":
            for i in range(1, 101):
                rams.append(float(ws.cell(row=2, column=i).value))
                cpus.append(float(ws.cell(row=2, column=i).value))
        elif last == "1":
            for i in range(2, 101):
                rams.append(float(ws.cell(row=2, column=i).value))
                cpus.append(float(ws.cell(row=2, column=i).value))
            rams.append(float(ws.cell(row=2, column=1).value))
            cpus.append(float(ws.cell(row=2, column=1).value))
        else:
            for i in range(int(last) + 1, 101):
                rams.append(float(ws.cell(row=2, column=i).value))
                cpus.append(float(ws.cell(row=2, column=i).value))
            for i in range(1, int(last) + 1):
                rams.append(float(ws.cell(row=2, column=i).value))
                cpus.append(float(ws.cell(row=2, column=i).value))
        plt.figure(figsize=(39, 18))
        ax1 = plt.subplot(1, 2, 1)
        ax1.set_ylim(0, 100)
        ax1.plot(list(range(1, 101)), cpus, 'b-')
        ax2 = plt.subplot(1, 2, 2)
        ax2.set_ylim(0, 100)
        ax2.plot(list(range(1, 101)), rams, 'b-')
        ax1.set_xlabel('Usage (%)', fontsize=44)
        ax1.set_ylabel('CPU')
        ax2.set_ylabel('RAM')
        ax1.xaxis.label.set_size(44)
        ax1.yaxis.label.set_size(44)
        ax2.yaxis.label.set_size(44)
        plt.savefig(str(ctx.author.id) + ".png", dpi=192)
        plt.clf()
        plt.close()
        wb.close()
        await ctx.send(file=discord.File("./" + str(ctx.author.id) + ".png"))
        os.remove(str(ctx.author.id) + '.png')

def setup(bot):
    bot.add_cog(Information_ko(bot))
